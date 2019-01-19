"""
The program processes schedule given in the text file and converts it into a spreadsheet (.xlsx file)
"""

import re
import datetime
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, Color, NamedStyle, PatternFill
from Schedule import Schedule


def import_file_contents(directory):
    txt = open(directory, "r", encoding='windows-1251')
    schedule_txt = txt.read()
    txt.close()
    return schedule_txt


def process_data(data):
    """The function processes text from the file and returns a list that is used to create a spreadsheet"""

    def get_header_info(header):
        group = research(r'(?<=Група )\w{2,3}(-\d{2})*', header)                    # (Група ){www-dd}
        year = research(r'\d{4}-\d{4} н.р.', header)                                # {dddd-dddd} н.р.
        semester = 1 if research(r'\w+(?= семестр)', header) is 'І' else 2          # I/II( семестр)
        first_monday = research(r'\d{2}.\d{2} ', header)[:-1]                       # dd.dd\

        start_date = datetime.date(int(year[0:4]) if semester == 1
                else int(year[5:9]), int(first_monday[3:]), int(first_monday[:2]))

        print('\nGroup: {}\nYear: {}\nSemester: {}\nFirst week date: {}'
              .format(group, year, semester, start_date))

        return group, year, semester, start_date

    def get_classes(schedule, day_desc_list):
        """The function analyzes data about periods and passes it to class Schedule"""
        year = schedule.start.year

        print('\nProcessing data...')

        # Loop through DAYS of week ====================================================================================
        for i in range(len(day_desc_list)):
            lessons_of_the_day = re.split(r'\n(?=\d пара)', day_desc_list[i])  # [day, num+disclist, num+disclist, ...]

            # Loop through LESSONS -------------------------------------------------------------------------------------
            for j in range(1, len(lessons_of_the_day)):
                courses_at_lesson = re.split('\n\* ', lessons_of_the_day[j])  # [num, disc1, disc2, .. ]
                # information about NUMBER of lesson has INDEX 0!!!
                lesson_number = int(research(r'\d(?= \w+)', courses_at_lesson[0]))

                # Loop through COURSES *********************************************************************************
                for crs in range(1, len(courses_at_lesson)):
                    # Getting the name of the COURSE
                    study_course = research(r'.+(?= \(\w\))', courses_at_lesson[crs])
                    # Excluding speciality from name of the course
                    study_course = re.sub(re.compile(r'\(.+\)'), '', study_course)
                    subgroup = None
                    # If a subgroup number is present, it is assigned to the variable and excluded from the name
                    if re.search(re.compile(r'\d'), study_course):
                        subgroup = research(r'\d', study_course)
                        study_course = re.sub(re.compile(r'\s*\d\s*'), '', study_course)
                        study_course = re.sub(re.compile(r'підгр?|підгрупа|група'), '', study_course)

                    lesson_type = research(r'(?<=\()\w(?=\))', courses_at_lesson[crs]) # (w)
                    teacher = research(r'(?<=\[).+(?=\])', courses_at_lesson[crs]) # [name]

                    # Getting room numbers and dates
                    rooms_list = re.findall(re.compile(r'ауд.\d{3} \(.{5,11}\)'), courses_at_lesson[crs])
                    for rm in rooms_list:
                        room = research(r'ауд.\d{3}', rm)

                        room_dates = research(r'(?<=\().{5,11}(?=\))', rm)
                        # Depending on a number of dates 1+ lessons can be added at once
                        if len(room_dates) == 5:
                            dt = datetime.date(year, int(room_dates[3:]), int(room_dates[:2]))
                            schedule.add_lesson(study_course, room, lesson_type, lesson_number, teacher, dt, subgroup)
                        else:
                            list_of_dates = []
                            first_date = datetime.date(year, int(room_dates[3:5]), int(room_dates[:2]))
                            last_date = datetime.date(year, int(room_dates[9:]), int(room_dates[6:8]))
                            shift = datetime.timedelta(days=7)
                            while first_date <= last_date:
                                list_of_dates.append(first_date)
                                first_date += shift
                                for dt in list_of_dates:
                                    schedule.add_lesson(study_course, room, lesson_type,
                                                        lesson_number, teacher, dt, subgroup)
        return schedule

    def research(regular, string):
        return re.search(re.compile(regular), string).group()

    # Splitting schedule into sections (0 - general info, 1-5 - days)
    divided_schedule = re.split('-{2,}\n', data)

    schedule_info = get_header_info(divided_schedule[0])
    scd = Schedule(schedule_info)
    # processing info about classes and passing it to Schedule
    scd = get_classes(scd, divided_schedule[1:])

    return scd


def create_spreadsheet(schedule):
    # CREATING WORKBOOK
    print('Creating spreadsheet...')
    filename = schedule.group + '.xlsx'
    s_wb = openpyxl.workbook.Workbook()
    sheet = s_wb.active
    sheet.title = schedule.group

    # ADDING INFORMATION
    # lessons
    days = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'П\'ятниця']
    sheet.append(days)
    lessons = schedule.create_spreadsheet()
    for row in lessons:
        sheet.append(row)
    # date + number
    sheet.insert_cols(0, 2)
    curr_date = schedule.start
    shift = datetime.timedelta(days=7)
    i = 2 # row number
    while sheet['C' + str(i)].value is not None:
        for j in range(0, 14, 2):
            sheet['A' + str(i + j)] = curr_date
            sheet['B' + str(i + j)] = (j + 2) // 2
        curr_date += shift
        i += 14
    sheet.insert_rows(0,2)
    # header
    sheet['A1'] = 'РОЗКЛАД НА {} СЕМЕСТР'.format(schedule.semester)
    sheet['A2'] = 'Група {}, {}'.format(schedule.group, schedule.year)

    # SETTING DIMENSIONS
    for i in range(2):
        sheet.column_dimensions[chr(ord('A') + i)].width = 4
    for i in range(5):
        sheet.column_dimensions[chr(ord('C') + i)].width = 19

    # MERGING TITLE CELLS
    sheet.merge_cells('A1:G1')
    sheet.merge_cells('A2:G2')
    sheet.merge_cells('A3:B3')

    # DELETING EMPTY ROWS
    i = 4
    while sheet['A' + str(i)].value is not None:
        if (sheet['C' + str(i)].value is ''
                and sheet['D' + str(i)].value is ''
                and sheet['E' + str(i)].value is ''
                and sheet['F' + str(i)].value is ''
                and sheet['G' + str(i)].value is ''):
            sheet.delete_rows(i, 2)
        else:
            i += 2

    # APPLYING STYLES
    s_wb = apply_styles(s_wb)

    # SAVING SPREADSHEET
    print('File {} is successfully saved'.format(filename))
    s_wb.save(filename)
    del s_wb
    del schedule


def apply_styles(workbook):

    def workbook_styles_init():
        # COLORS
        blk = Color('000000')
        ttl = Color('992600')
        col_styles = (Color('ffffcc'), Color('e6ffcc'))

        # ALIGNMENT
        al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        al_cent_rot = Alignment(horizontal='center', vertical='center', wrap_text=False, text_rotation=90)
        al_bottom = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        al_top = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # BORDERS
        side = Side(color=blk, style='thin')
        all_borders = Border(top=side, bottom=side, left=side, right=side)
        without_bot = Border(top=side, left=side, right=side)
        without_top = Border(bottom=side, left=side, right=side)

        # FONTS
        f_title = Font(name='Calibri Light', size=12, bold=True, color=ttl)
        f_subtitle = Font(name='Calibri Light', size=10, color=ttl)
        f_day_date = Font(name='Calibri Light', size=10)
        f_disc = Font(name='Calibri', size=11)
        f_room = Font(name='Calibri', size=9)

        # STYLES
        s_title = NamedStyle(name='title', font=f_title, alignment=al_bottom, border=without_bot)
        s_sub = NamedStyle(name='subtitle', font=f_subtitle, alignment=al_top, border=without_top)
        s_day = NamedStyle(name='day', font=f_day_date, alignment=al_center, border=all_borders)
        s_date = NamedStyle(name='date', font=f_day_date, alignment=al_cent_rot, border=all_borders)
        s_disc = NamedStyle(name='disc', font=f_disc, alignment=al_center, border=without_bot)
        s_room = NamedStyle(name='room', font=f_room, alignment=al_top, border=without_top)
        s_num = NamedStyle(name='snum', font=f_day_date, alignment=al_center, border=all_borders)

        return s_title, s_sub, s_day, s_date, s_disc, s_room, s_num, col_styles

    def style_week(sheet, start, i, disc, room, num, date, fill):
        sheet.merge_cells('A{}:A{}'.format(start, i + 1))
        for j in range(start, i + 2, 2):
            # Lessons cells
            for k in sheet[j]:
                k.style = disc
                k.fill = fill
            for k in sheet[j + 1]:
                k.style = room
                k.fill = fill

            # Date cells styling
            sheet['A' + str(j)].style = date
            sheet['A' + str(j)].number_format = 'DD.MM'
            sheet['A' + str(j + 1)].style = date

            # Number cells styling
            sheet.merge_cells('B{}:B{}'.format(j, j + 1))
            sheet['B' + str(j)].style = num
            sheet['B' + str(j + 1)].style = num

        return sheet

    sheet = workbook.active

    # INITIALIZING STYLES
    title, subtitle, day, date, disc, room, num, col_st = workbook_styles_init()
    workbook.add_named_style(title)
    workbook.add_named_style(subtitle)
    workbook.add_named_style(day)
    workbook.add_named_style(date)
    workbook.add_named_style(disc)
    workbook.add_named_style(room)
    workbook.add_named_style(num)
    col1, col2 = col_st
    # fills
    fill1 = PatternFill(fgColor=col1, fill_type='solid')
    fill2 = PatternFill(fgColor=col2, fill_type='solid')

    # APPLYING STYLES
    # head section
    for i in sheet[1]:
        i.style = title
    for i in sheet[2]:
        i.style = subtitle
    for i in sheet[3]:
        i.style = day

    # lessons section
    i = 4
    start = 4
    style = 0
    while sheet['B' + str(i)].value is not None:
        fill = fill1 if style % 2 == 0 else fill2
        if sheet['B' + str(i + 2)].value is None:
            sheet = style_week(sheet, start, i, disc, room, num, date, fill)
        elif sheet['B' + str(i + 2)].value < sheet['B' + str(i)].value:
            sheet = style_week(sheet, start, i, disc, room, num, date, fill)
            start = i + 2
            style += 1
        i += 2
    return workbook


def main():
    while True:
        """Loop requests a name of the file until either the correct
        path is entered or the program is terminated by a user entering n."""

        # dir = input("Enter the file name (path): ")
        dir = 'IoT-11-2017-2.txt'

        try:
            file_text = import_file_contents(dir)
        except FileNotFoundError:
            print("Oops! It seems like there is no such file.")
            ans = input("Do you want to try again? (Y)es or (N)o? ")
            if ans.lower() == 'y':
                continue
        else:
            schedule = process_data(file_text)
            create_spreadsheet(schedule)
        break


if __name__ == "__main__":
    main()
