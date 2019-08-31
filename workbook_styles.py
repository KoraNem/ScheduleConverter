from openpyxl.styles import Border, Side, Alignment, Font, Color, NamedStyle, PatternFill


def workbook_styles_init():
    # COLORS
    blk = Color('000000')
    ttl = Color('992600')
    col_styles = (Color('ffffcc'), Color('e6ffcc'))

    # ALIGNMENT
    al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_cent_rot = Alignment(horizontal='center', vertical='center', wrap_text=False, text_rotation=90)
    al_bottom = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
    al_top = Alignment(horizontal='center', vertical='top', wrap_text=True)

    # BORDERS
    side = Side(color=blk, style='thin')
    all_borders = Border(top=side, bottom=side, left=side, right=side)
    without_bot = Border(top=side, left=side, right=side)
    without_top = Border(bottom=side, left=side, right=side)

    # FONTS
    f_title = Font(name='Calibri Light', size=12, bold=True, color=ttl)
    f_subtitle = Font(name='Calibri Light', size=10, color=ttl)
    f_day_date = Font(name='Calibri Light', size=10)
    f_disc = Font(name='Calibri', size=11)
    f_room = Font(name='Calibri', size=9)

    # STYLES
    s_title = NamedStyle(name='title', font=f_title, alignment=al_bottom, border=without_bot)
    s_subtitle = NamedStyle(name='subtitle', font=f_subtitle, alignment=al_top, border=without_top)
    s_day = NamedStyle(name='day', font=f_day_date, alignment=al_center, border=all_borders)
    s_date = NamedStyle(name='date', font=f_day_date, alignment=al_cent_rot, border=all_borders)
    s_lesson = NamedStyle(name='disc', font=f_disc, alignment=al_center, border=without_bot)
    s_room = NamedStyle(name='room', font=f_room, alignment=al_top, border=without_top)
    s_num = NamedStyle(name='snum', font=f_day_date, alignment=al_center, border=all_borders)

    return s_title, s_subtitle, s_day, s_date, s_lesson, s_room, s_num, col_styles


def style_week(sheet, start, i, lesson, room, num, date, fill):
    sheet.merge_cells('A{}:A{}'.format(start, i + 1))
    for j in range(start, i + 2, 2):
        # Lessons cells
        for k in sheet[j]:
            k.style = lesson
            k.fill = fill
        for k in sheet[j + 1]:
            k.style = room
            k.fill = fill

        # Date cells styling
        sheet['A' + str(j)].style = date
        sheet['A' + str(j)].number_format = 'DD.MM'
        sheet['A' + str(j + 1)].style = date

        # Number cells styling
        sheet.merge_cells('B{}:B{}'.format(j, j + 1))
        sheet['B' + str(j)].style = num
        sheet['B' + str(j + 1)].style = num

    return sheet


def apply_styles(workbook):
    sheet = workbook.active

    # INITIALIZING STYLES
    title, subtitle, day, date, lesson, room, num, col_st = workbook_styles_init()
    workbook.add_named_style(title)
    workbook.add_named_style(subtitle)
    workbook.add_named_style(day)
    workbook.add_named_style(date)
    workbook.add_named_style(lesson)
    workbook.add_named_style(room)
    workbook.add_named_style(num)
    col1, col2 = col_st
    # fills
    fill1 = PatternFill(fgColor=col1, fill_type='solid')
    fill2 = PatternFill(fgColor=col2, fill_type='solid')

    # APPLYING STYLES
    # header section
    for i in sheet[1]:
        i.style = title
    for i in sheet[2]:
        i.style = subtitle
    for i in sheet[3]:
        i.style = day

    # lessons section
    i = 4
    start = 4
    style = 0
    while sheet['B' + str(i)].value is not None:
        fill = fill1 if style % 2 == 0 else fill2
        if sheet['B' + str(i + 2)].value is None:
            sheet = style_week(sheet, start, i, lesson, room, num, date, fill)
        elif sheet['B' + str(i + 2)].value < sheet['B' + str(i)].value:
            sheet = style_week(sheet, start, i, lesson, room, num, date, fill)
            start = i + 2
            style += 1
        i += 2
    return workbook
