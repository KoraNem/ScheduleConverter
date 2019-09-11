"""
The program processes schedule given in the text file and converts it into a spreadsheet (.xlsx file)
"""
from os import remove
import datetime
import openpyxl
from processing_input import process_data
from workbook_styles import apply_styles
from pdf_to_txt import pdf_txt


def import_file_contents(directory, filename):
    if filename[-3:] == "pdf":
        pdf_txt(directory + '\\' + filename)
        txt = open(filename[:-3] + "txt", "r")
    else:
        txt = open(directory + '\\' + filename, "r")
    schedule_txt = txt.read()
    txt.close()
    if filename[-3:] == "pdf":
        remove(filename[:-3] + "txt")
    return schedule_txt


def create_spreadsheet(schedule):
    # CREATING WORKBOOK
    print('Creating spreadsheet...')
    filename = schedule.group + '.xlsx'
    s_wb = openpyxl.workbook.Workbook()
    sheet = s_wb.active
    sheet.title = schedule.group

    # ADDING INFORMATION
    # lessons
    days = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'П\'ятниця']
    sheet.append(days)
    lessons = schedule.create_spreadsheet()
    for row in lessons:
        sheet.append(row)
    # date + number
    sheet.insert_cols(0, 2)
    curr_date = schedule.start
    shift = datetime.timedelta(days=7)
    i = 2  # row number
    while sheet['C' + str(i)].value is not None:
        for j in range(0, 14, 2):
            sheet['A' + str(i + j)] = curr_date
            sheet['B' + str(i + j)] = (j + 2) // 2
        curr_date += shift
        i += 14
    sheet.insert_rows(0, 2)
    # header
    sheet['A1'] = 'РОЗКЛАД НА {} СЕМЕСТР'.format(schedule.semester)
    sheet['A2'] = 'Група {}, {}'.format(schedule.group, schedule.year)

    # SETTING DIMENSIONS
    for i in range(2):
        sheet.column_dimensions[chr(ord('A') + i)].width = 4
    for i in range(5):
        sheet.column_dimensions[chr(ord('C') + i)].width = 19

    # MERGING TITLE CELLS
    sheet.merge_cells('A1:G1')
    sheet.merge_cells('A2:G2')
    sheet.merge_cells('A3:B3')

    # DELETING EMPTY ROWS
    i = 4
    while sheet['A' + str(i)].value is not None:
        if (sheet['C' + str(i)].value is ''
                and sheet['D' + str(i)].value is ''
                and sheet['E' + str(i)].value is ''
                and sheet['F' + str(i)].value is ''
                and sheet['G' + str(i)].value is ''):
            sheet.delete_rows(i, 2)
        else:
            i += 2

    # APPLYING STYLES
    s_wb = apply_styles(s_wb)

    # SAVING SPREADSHEET
    while True:
        try:
            s_wb.save(filename)
        except PermissionError:
            print('Please, close the existing {} file to save a new one!!!'.format(filename), end='')
            input()
        else:
            print('File {} is successfully saved'.format(filename))
            break

    del s_wb
    del schedule


def main():
    while True:
        """Loop requests a name of the file until either the correct
        path is entered or the program is terminated by a user entering n."""

        directory = input("Enter the path to directory: ")
        filename = input("Enter the file name: ")
        try:
            file_text = import_file_contents(directory, filename)
        except FileNotFoundError:
            print("Oops! It seems like there is no such file.")
            ans = input("Do you want to try again? (Y)es or (N)o? ")
            if ans.lower() == 'y':
                continue
        else:
            schedule = process_data(file_text)
            create_spreadsheet(schedule)
        break


if __name__ == "__main__":
    main()
